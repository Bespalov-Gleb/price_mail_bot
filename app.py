import asyncio
import contextlib
import logging
import os
import random
import smtplib
import ssl
import shutil
import tempfile
from email.message import EmailMessage
from pathlib import Path

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.filters import Command
from aiogram.types import Message
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("price_mail_bot")

BOT_TOKEN = os.getenv("PRICE_BOT_TOKEN", "")
PROXY = os.getenv("PRICE_BOT_PROXY") or None

ALLOWED_IDS = {
    int(x.strip())
    for x in (os.getenv("PRICE_BOT_ALLOWED_IDS", "") or "").split(",")
    if x.strip().isdigit()
}

PRICE_FILE_PATH = Path(os.getenv("PRICE_FILE_PATH", "./data/price.xlsx")).resolve()
PRICE_STATE_PATH = Path(os.getenv("PRICE_STATE_PATH", "./data/price_state.txt")).resolve()
PRICE_SHEET_NAME = (os.getenv("PRICE_SHEET_NAME") or "").strip()
PRICE_MAIL_INTERVAL_SECONDS = int(os.getenv("PRICE_MAIL_INTERVAL_SECONDS", "3600"))

SMTP_HOST = os.getenv("YANDEX_SMTP_HOST", "smtp.yandex.ru")
SMTP_PORT = int(os.getenv("YANDEX_SMTP_PORT", "465"))
SMTP_LOGIN = os.getenv("YANDEX_SMTP_LOGIN", "")
SMTP_PASSWORD = os.getenv("YANDEX_SMTP_PASSWORD", "")
EMAIL_TO = os.getenv("YANDEX_EMAIL_TO", "")
EMAIL_SUBJECT = os.getenv("YANDEX_EMAIL_SUBJECT", "Прайс-лист")

router = Router()
file_lock = asyncio.Lock()


def _create_bot() -> Bot:
    if PROXY:
        return Bot(token=BOT_TOKEN, session=AiohttpSession(proxy=PROXY))
    return Bot(token=BOT_TOKEN)


bot = _create_bot()


def _allowed(user_id: int) -> bool:
    if not ALLOWED_IDS:
        return True
    return user_id in ALLOWED_IDS


def _read_direction() -> int:
    # +1 if missing
    try:
        raw = PRICE_STATE_PATH.read_text(encoding="utf-8").strip()
        return -1 if raw == "-1" else 1
    except Exception:
        return 1


def _write_next_direction(current: int) -> None:
    nxt = -1 if current == 1 else 1
    PRICE_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRICE_STATE_PATH.write_text(str(nxt), encoding="utf-8")


def _collect_numeric_cells(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(path)
    try:
        sheets = [wb[PRICE_SHEET_NAME]] if PRICE_SHEET_NAME and PRICE_SHEET_NAME in wb.sheetnames else wb.worksheets
        cells: list[tuple[str, str]] = []
        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and float(cell.value) > 0:
                        cells.append((ws.title, cell.coordinate))
        return cells
    finally:
        wb.close()


def _mutate_price() -> tuple[bool, str]:
    if not PRICE_FILE_PATH.exists():
        return False, f"Файл не найден: {PRICE_FILE_PATH}"

    cells = _collect_numeric_cells(PRICE_FILE_PATH)
    if not cells:
        return False, "Нет числовых ячеек > 0 для изменения"

    direction = _read_direction()
    sheet_name, coord = random.choice(cells)

    wb = load_workbook(PRICE_FILE_PATH)
    try:
        ws = wb[sheet_name]
        old = float(ws[coord].value)
        new = old + float(direction)  # +/- 1 RUB
        if new < 0:
            new = 0.0
        ws[coord].value = new

        PRICE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            dir=str(PRICE_FILE_PATH.parent),
        )
        tmp.close()
        wb.save(tmp.name)
        os.replace(tmp.name, str(PRICE_FILE_PATH))
    finally:
        wb.close()

    _write_next_direction(direction)
    sign = "+" if direction == 1 else "-"
    return True, f"{sheet_name}!{coord}: {old} -> {new} ({sign}1 RUB)"


def _send_mail(file_path: Path) -> None:
    if not SMTP_LOGIN or not SMTP_PASSWORD or not EMAIL_TO:
        raise RuntimeError("Не заполнены SMTP параметры Yandex в .env")

    msg = EmailMessage()
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"] = SMTP_LOGIN
    msg["To"] = EMAIL_TO
    msg.set_content("Обновленный прайс во вложении.")

    content = file_path.read_bytes()
    msg.add_attachment(
        content,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file_path.name,
    )

    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as s:
        s.login(SMTP_LOGIN, SMTP_PASSWORD)
        s.send_message(msg)


async def _tick() -> None:
    async with file_lock:
        ok, info = _mutate_price()
        if not ok:
            logger.warning("Tick skipped: %s", info)
            return
        logger.info("Price changed: %s", info)
        _send_mail(PRICE_FILE_PATH)
        logger.info("Mail sent to %s", EMAIL_TO)


async def scheduler_loop() -> None:
    interval = max(30, PRICE_MAIL_INTERVAL_SECONDS)
    logger.info("Scheduler started, interval=%s sec", interval)
    while True:
        try:
            await _tick()
        except Exception as e:
            logger.exception("Scheduler error: %s", e)
        await asyncio.sleep(interval)


@router.message(Command("start"))
async def cmd_start(message: Message):
    if not _allowed(message.from_user.id):
        await message.answer("⛔ Доступ запрещен")
        return
    await message.answer(
        "Готово.\n"
        "1) По таймеру меняю случайную цену на +/-1 RUB и отправляю файл на почту.\n"
        "2) Пришли .xlsx файлом — обновлю серверный файл без сброса таймера."
    )


@router.message(Command("status"))
async def cmd_status(message: Message):
    if not _allowed(message.from_user.id):
        await message.answer("⛔ Доступ запрещен")
        return
    await message.answer(
        f"Файл: {PRICE_FILE_PATH}\n"
        f"Есть: {'да' if PRICE_FILE_PATH.exists() else 'нет'}\n"
        f"Интервал: {max(30, PRICE_MAIL_INTERVAL_SECONDS)} сек\n"
        f"Почта: {EMAIL_TO or '-'}"
    )


@router.message(F.document)
async def on_document(message: Message):
    if not _allowed(message.from_user.id):
        await message.answer("⛔ Доступ запрещен")
        return

    doc = message.document
    file_name = (doc.file_name or "").lower()
    if not file_name.endswith(".xlsx"):
        await message.answer("Нужен файл .xlsx")
        return

    async with file_lock:
        PRICE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            dir=str(PRICE_FILE_PATH.parent),
        )
        tmp.close()
        try:
            await bot.download(doc, destination=tmp.name)
            wb = load_workbook(tmp.name)
            wb.close()
            os.replace(tmp.name, str(PRICE_FILE_PATH))
        except Exception as e:
            # fallback на случай edge-кейсов с разными FS
            with contextlib.suppress(Exception):
                shutil.move(tmp.name, str(PRICE_FILE_PATH))
                await message.answer("✅ Файл обновлен. Таймер не сброшен.")
                return
            with contextlib.suppress(Exception):
                os.unlink(tmp.name)
            await message.answer(f"Ошибка загрузки/валидации: {e}")
            return

    await message.answer("✅ Файл обновлен. Таймер не сброшен.")


async def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("PRICE_BOT_TOKEN не задан в .env")

    dp = Dispatcher()
    dp.include_router(router)

    task = asyncio.create_task(scheduler_loop())
    try:
        await dp.start_polling(bot)
    finally:
        task.cancel()
        with contextlib.suppress(Exception):
            await task


if __name__ == "__main__":
    asyncio.run(main())
